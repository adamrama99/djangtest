import io
import json
import os
from functools import wraps
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse, HttpResponseForbidden
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import BooleanField, Case, Q, Value, When
from django.template.loader import render_to_string
from django.utils import timezone
from .models import (
    DocumentationRequest, LEDType, Requirement, ViewPhoto, cameratype,
    BrandMateri, Lokasi, Dokumentator, DocumentationRequestLokasiAssignment, EditHistory,
    DocumentationRequestMateri,
    MaintenanceRequest, MaintenanceRequestMateri,
    JadwalTayang, JadwalTayangMateri, JadwalTayangFotoTayang, JadwalTayangBuktiPlaylist, JadwalTayangFotoTakeout,
    TakeoutAlertRule,
    UserLoginRequirement,
)
from .forms import (
    DocumentationRequestForm,
    DocumentationRequestEditForm,
    DocumentationRequestProofForm,
    MasterDataForm,
    MaintenanceRequestForm,
    MaintenanceRequestEditForm,
    MaintenanceRequestProofForm,
    JadwalTayangForm,
    JadwalTayangEditForm,
    TakeoutAlertRuleForm,
    UserForm,
)
from .notifications import get_active_takeout_notifications
from django.core.paginator import Paginator
from django.contrib.auth import get_user_model

User = get_user_model()


def _is_ajax(request):
    return request.headers.get("X-Requested-With") == "XMLHttpRequest"


def _load_openpyxl():
    try:
        import openpyxl
    except ImportError:
        return None
    return openpyxl


def _is_admin(user):
    """Check if user is in the 'admin' group or is a superuser."""
    return user.is_superuser or user.groups.filter(name="admin").exists()


def _is_requester(user):
    """Check if user is in the 'requester' group."""
    return user.groups.filter(name="requester").exists()


def _is_staff_role(user):
    """Check if user is in the 'staff' group."""
    return user.groups.filter(name="staff").exists()


def _is_executor(user):
    """Check if user is in the 'executor' group."""
    return user.groups.filter(name="executor").exists()


def _can_view_all_service_requests(user):
    """Admins and staff can monitor all service requests."""
    return _is_admin(user) or _is_staff_role(user)


def _can_access_request_edit(user):
    """Admins, staff, and requesters can access edit pages."""
    return _is_admin(user) or _is_staff_role(user) or _is_requester(user)


def _can_manage_service_pelaksana(user):
    """Admins and staff can assign pelaksana/dokumentator."""
    return _is_admin(user) or _is_staff_role(user)


def _can_upload_service_proof(user):
    """Admins, staff, and executors can upload a single proof photo."""
    return _is_admin(user) or _is_staff_role(user) or _is_executor(user)


def _user_dokumentator_candidate_names(user):
    candidate_names = []
    full_name = user.get_full_name().strip()
    if full_name:
        candidate_names.append(full_name)
    for value in (user.first_name, user.last_name):
        cleaned = (value or "").strip()
        if cleaned:
            candidate_names.append(cleaned)
    username = (user.username or "").strip()
    if username:
        if "@" in username:
            local_part = username.split("@", 1)[0].strip()
            if local_part:
                candidate_names.append(local_part)
        candidate_names.append(username)

    seen = set()
    unique_names = []
    for name in candidate_names:
        normalized = name.casefold()
        if normalized not in seen:
            seen.add(normalized)
            unique_names.append(name)
    return unique_names


def _user_dokumentator_display_name(user):
    candidate_names = _user_dokumentator_candidate_names(user)
    return candidate_names[0] if candidate_names else ""


def _matching_dokumentator_queryset(user):
    candidate_names = _user_dokumentator_candidate_names(user)
    if not candidate_names:
        return Dokumentator.objects.none()

    query = Q(pk__isnull=True)
    for name in candidate_names:
        query |= Q(name__iexact=name)
    return Dokumentator.objects.filter(query)


def _matching_dokumentator_ids(user):
    return list(_matching_dokumentator_queryset(user).values_list("id", flat=True))


def _assignable_dokumentators_queryset():
    """
    Keep Dokumentator choices in sync with executor users so select2 options
    reflect the current executor roster.
    """
    executor_users = (
        User.objects.filter(groups__name="executor")
        .order_by("first_name", "last_name", "username")
        .distinct()
    )

    for executor_user in executor_users:
        display_name = _user_dokumentator_display_name(executor_user)
        if not display_name:
            continue
        if not _matching_dokumentator_queryset(executor_user).exists():
            Dokumentator.objects.get_or_create(name=display_name)

    return Dokumentator.objects.all().order_by("name")


def _doc_request_label(doc_request):
    brand = doc_request.brand.name if doc_request.brand else "N/A"
    label = f"{brand} - {doc_request.tanggal}"
    if not getattr(doc_request, "pk", None):
        return label

    try:
        lokasi_label = doc_request.lokasi_display()
    except ValueError:
        lokasi_label = "-"

    if lokasi_label and lokasi_label != "-":
        return f"{label} - {lokasi_label}"
    return label


def _jadwal_tayang_label(jadwal_tayang):
    brand = jadwal_tayang.brand.name if jadwal_tayang.brand else "N/A"
    label = f"{brand} - {jadwal_tayang.tanggal_tayang}"
    if not getattr(jadwal_tayang, "pk", None):
        return label

    lokasi_label = jadwal_tayang.lokasi_display()
    if lokasi_label and lokasi_label != "-":
        return f"{label} - {lokasi_label}"
    return label


def _joined_names(queryset, empty_label="Belum ditentukan"):
    names = list(queryset.order_by("name").values_list("name", flat=True))
    return ", ".join(names) if names else empty_label


def _get_search_query(request):
    return request.GET.get("q", "").strip()


def _search_context(request, placeholder):
    params = request.GET.copy()
    params.pop("page", None)
    search_query = _get_search_query(request)
    return {
        "search_query": search_query,
        "search_active": bool(search_query),
        "search_placeholder": placeholder,
        "page_query": params.urlencode(),
    }


def _pk_search_q(search_query):
    if search_query.isdigit():
        return Q(pk=int(search_query))
    return Q(pk__isnull=True)


def _group_jadwal_tayang_by_lokasi(jadwal_list):
    grouped = {}

    for jadwal_tayang in jadwal_list:
        lokasi_list = list(jadwal_tayang.lokasi.all())
        if not lokasi_list:
            grouped.setdefault("Tanpa Lokasi", {"lokasi_name": "Tanpa Lokasi", "items": []})["items"].append(jadwal_tayang)
            continue

        for lokasi in sorted(lokasi_list, key=lambda item: item.name.casefold()):
            grouped.setdefault(
                lokasi.name,
                {"lokasi_name": lokasi.name, "items": []},
            )["items"].append(jadwal_tayang)

    return [grouped[key] for key in sorted(grouped.keys(), key=str.casefold)]


def _jadwal_tayang_photo_status_info(jadwal_tayang, now=None):
    if now is None:
        now = timezone.now()

    materi_items = list(jadwal_tayang.materi_items.all())
    total_materi = len(materi_items)
    if not total_materi:
        return {
            "label": "Belum Ada Materi",
            "badge_class": "secondary",
            "detail": "Tambahkan materi sebelum executor upload foto.",
        }

    foto_tayang_count = sum(1 for materi in materi_items if materi.foto_tayang_set.exists())
    foto_takeout_count = sum(1 for materi in materi_items if materi.foto_takeout_set.exists())
    bukti_playlist_count = sum(1 for materi in materi_items if _materi_has_bukti_playlist(materi))

    if foto_takeout_count == total_materi:
        return {
            "label": "Sudah Upload Foto Takeout",
            "badge_class": "success",
            "detail": f"{foto_takeout_count}/{total_materi} materi sudah punya foto takeout.",
        }

    if now > jadwal_tayang.tanggal_takeout:
        return {
            "label": "Belum Takeout",
            "badge_class": "danger",
            "detail": f"Waktu takeout sudah lewat. Foto takeout {foto_takeout_count}/{total_materi} materi.",
        }

    if foto_tayang_count == total_materi and bukti_playlist_count == total_materi:
        return {
            "label": "Sudah Upload Foto Tayang + Playlist",
            "badge_class": "primary",
            "detail": f"Foto tayang dan playlist lengkap untuk {total_materi} materi.",
        }

    if foto_tayang_count or bukti_playlist_count:
        return {
            "label": "Upload Foto Parsial",
            "badge_class": "info",
            "detail": f"Foto tayang {foto_tayang_count}/{total_materi}, playlist {bukti_playlist_count}/{total_materi}.",
        }

    return {
        "label": "Belum Upload Foto",
        "badge_class": "secondary",
        "detail": f"Belum ada foto untuk {total_materi} materi.",
    }


def _materi_has_bukti_playlist(materi):
    try:
        bukti = materi.bukti_playlist
    except JadwalTayangBuktiPlaylist.DoesNotExist:
        return False
    return bool(bukti.foto_pagi or bukti.foto_siang or bukti.foto_malam)


def _extract_materi_rows(post_data):
    materi_ids = post_data.getlist("materi_ids[]")
    materi_names = post_data.getlist("materi_names[]")
    max_len = max(len(materi_ids), len(materi_names))
    rows = []

    for index in range(max_len):
        raw_id = materi_ids[index].strip() if index < len(materi_ids) else ""
        nama_materi = materi_names[index].strip() if index < len(materi_names) else ""
        if not nama_materi:
            continue
        rows.append({
            "id": raw_id if raw_id.isdigit() else "",
            "nama_materi": nama_materi,
        })

    return rows


def _initial_materi_rows(jadwal_tayang=None):
    if jadwal_tayang and jadwal_tayang.pk:
        rows = [
            {"id": str(materi.id), "nama_materi": materi.nama_materi}
            for materi in jadwal_tayang.materi_items.order_by("sort_order", "id")
        ]
        if rows:
            return rows
    return [{"id": "", "nama_materi": ""}]


def _sync_jadwal_tayang_materi(jadwal_tayang, materi_rows):
    existing = {
        str(materi.id): materi
        for materi in jadwal_tayang.materi_items.all()
    }
    keep_ids = []

    for index, row in enumerate(materi_rows):
        materi = existing.get(row["id"])
        if materi:
            if materi.nama_materi != row["nama_materi"] or materi.sort_order != index:
                materi.nama_materi = row["nama_materi"]
                materi.sort_order = index
                materi.save(update_fields=["nama_materi", "sort_order"])
        else:
            materi = JadwalTayangMateri.objects.create(
                jadwal_tayang=jadwal_tayang,
                nama_materi=row["nama_materi"],
                sort_order=index,
            )
        keep_ids.append(materi.id)

    jadwal_tayang.materi_items.exclude(id__in=keep_ids).delete()
    jadwal_tayang.auto_update_status()


def _sync_documentation_request_materi(doc_request, materi_rows):
    existing = {
        str(materi.id): materi
        for materi in doc_request.materi_items.all()
    }
    keep_ids = []

    for index, row in enumerate(materi_rows):
        materi = existing.get(row["id"])
        if materi:
            if materi.nama_materi != row["nama_materi"] or materi.sort_order != index:
                materi.nama_materi = row["nama_materi"]
                materi.sort_order = index
                materi.save(update_fields=["nama_materi", "sort_order"])
        else:
            materi = DocumentationRequestMateri.objects.create(
                documentation_request=doc_request,
                nama_materi=row["nama_materi"],
                sort_order=index,
            )
        keep_ids.append(materi.id)

    doc_request.materi_items.exclude(id__in=keep_ids).delete()


def _sync_maintenance_request_materi(maint_request, materi_rows):
    existing = {
        str(materi.id): materi
        for materi in maint_request.materi_items.all()
    }
    keep_ids = []

    for index, row in enumerate(materi_rows):
        materi = existing.get(row["id"])
        if materi:
            if materi.nama_materi != row["nama_materi"] or materi.sort_order != index:
                materi.nama_materi = row["nama_materi"]
                materi.sort_order = index
                materi.save(update_fields=["nama_materi", "sort_order"])
        else:
            materi = MaintenanceRequestMateri.objects.create(
                maintenance_request=maint_request,
                nama_materi=row["nama_materi"],
                sort_order=index,
            )
        keep_ids.append(materi.id)

    maint_request.materi_items.exclude(id__in=keep_ids).delete()


def _materi_display_from_rows(materi_rows):
    names = [row["nama_materi"] for row in materi_rows]
    return ", ".join(names) if names else "-"


def _contains_search_value(search_query, *values):
    normalized_query = search_query.casefold()
    for value in values:
        if value is None:
            continue
        if normalized_query in str(value).casefold():
            return True
    return False


def _annotate_login_required_flag(queryset):
    return queryset.annotate(
        login_required_flag=Case(
            When(login_requirement__requires_login=False, then=Value(False)),
            default=Value(True),
            output_field=BooleanField(),
        )
    )


def _user_search_filter(search_query):
    return (
        _pk_search_q(search_query)
        | Q(username__icontains=search_query)
        | Q(first_name__icontains=search_query)
        | Q(last_name__icontains=search_query)
        | Q(email__icontains=search_query)
        | Q(groups__name__icontains=search_query)
    )


def _get_or_create_dokumentator_for_user(user):
    candidate_names = _user_dokumentator_candidate_names(user)
    dokumentator = _matching_dokumentator_queryset(user).order_by("name").first()
    if dokumentator:
        return dokumentator, False

    primary_name = candidate_names[0] if candidate_names else ""
    if not primary_name:
        return None, False

    dokumentator, created = Dokumentator.objects.get_or_create(name=primary_name)
    return dokumentator, created


def _create_edit_history(
    *,
    user,
    action,
    request_type,
    object_id,
    label,
    field_name="",
    old_value="",
    new_value="",
):
    EditHistory.objects.create(
        user=user,
        action=action,
        request_type=request_type,
        doc_request_id=object_id,
        doc_request_label=label,
        field_name=field_name,
        old_value=old_value,
        new_value=new_value,
    )


def _filter_doc_requests_for_user(queryset, user):
    if _can_view_all_service_requests(user):
        return queryset
    if _is_executor(user):
        dokumentator_ids = _matching_dokumentator_ids(user)
        if not dokumentator_ids:
            return queryset.none()
        return queryset.filter(lokasi_assignments__pelaksana__in=dokumentator_ids).distinct()
    return queryset.filter(submitted_by=user)


def _filter_maint_requests_for_user(queryset, user):
    if _can_view_all_service_requests(user):
        return queryset
    if _is_executor(user):
        dokumentator_ids = _matching_dokumentator_ids(user)
        if not dokumentator_ids:
            return queryset.none()
        return queryset.filter(pelaksana__in=dokumentator_ids).distinct()
    return queryset.filter(submitted_by=user)


def _can_edit_request_record(user, submitted_by):
    if not _can_access_request_edit(user):
        return False
    if _is_admin(user) or _is_staff_role(user):
        return True
    submitted_by_id = submitted_by.pk if hasattr(submitted_by, "pk") else submitted_by
    return submitted_by_id == user.id


def _format_datetime_for_history(value):
    if not value:
        return "-"
    return timezone.localtime(value).strftime("%d/%m/%Y %H:%M")


def _format_file_for_history(field_file):
    if not field_file or not getattr(field_file, "name", ""):
        return "-"
    return os.path.basename(field_file.name)


def _validate_uploaded_image_size(uploaded_file):
    if uploaded_file and hasattr(uploaded_file, "size") and uploaded_file.size > 10 * 1024 * 1024:
        return "Ukuran file maksimal 10 MB."
    return ""


def _proof_status_from_materi(materi_items):
    total_materi = len(materi_items)
    if not total_materi:
        return "TODO"
    proof_count = sum(1 for materi in materi_items if materi.foto_bukti_kerja)
    if proof_count == total_materi:
        return "DONE"
    if proof_count:
        return "IN_PROGRESS"
    return "TODO"


def _sync_parent_status_from_materi_proofs(request_obj, materi_items):
    new_status = _proof_status_from_materi(materi_items)
    if request_obj.status != new_status:
        request_obj.status = new_status
        request_obj.save(update_fields=["status"])
    return new_status


def _doc_request_edit_snapshot(doc_request):
    return {
        "Brand": doc_request.brand.name if doc_request.brand else "-",
        "Lokasi": doc_request.lokasi_display(),
        "Materi": doc_request.materi_display(),
        "Jenis Produk": doc_request.jenis_led.name if doc_request.jenis_led else "-",
        "Tanggal": doc_request.tanggal.strftime("%d/%m/%Y") if doc_request.tanggal else "-",
        "Requirements": _joined_names(doc_request.requirements, empty_label="-"),
        "View Photo": _joined_names(doc_request.view_photo, empty_label="-"),
        "Jenis Kamera": _joined_names(doc_request.jenis_kamera, empty_label="-"),
        "PIC Pemohon": doc_request.pic_pemohon or "-",
        "Note": doc_request.note or "-",
    }


def _jadwal_tayang_edit_snapshot(jadwal_tayang):
    return {
        "Brand": jadwal_tayang.brand.name if jadwal_tayang.brand else "-",
        "Lokasi": jadwal_tayang.lokasi_display(),
        "Materi": jadwal_tayang.materi_display(),
        "Jenis Produk": jadwal_tayang.jenis_led.name if jadwal_tayang.jenis_led else "-",
        "Tanggal Tayang": _format_datetime_for_history(jadwal_tayang.tanggal_tayang),
        "Tanggal Takeout": _format_datetime_for_history(jadwal_tayang.tanggal_takeout),
        "PIC Pemohon": jadwal_tayang.pic_pemohon or "-",
        "Notes Requester": jadwal_tayang.note_requester or "-",
    }


def _serialize_notification_for_json(notification):
    return {
        "key": notification["key"],
        "title": notification["title"],
        "message": notification["message"],
        "detail_url": notification["detail_url"],
        "urgency": notification["urgency"],
        "urgency_label": notification["urgency_label"],
        "takeout_at_display": notification["takeout_at_display"],
        "offset_display": notification["offset_display"],
        "time_status": notification["time_status"],
        "pending_materi_label": notification["pending_materi_label"],
        "progress_label": notification["progress_label"],
    }


def _forbidden_response(request, message="Anda tidak memiliki izin untuk mengakses halaman ini."):
    """Render a proper 403 page with back button."""
    from django.template.response import TemplateResponse
    response = TemplateResponse(request, "products/403.html", {"message": message}, status=403)
    return response


def admin_required(view_func):
    """Decorator that restricts access to admin group only."""
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if not _is_admin(request.user):
            return _forbidden_response(request, "Halaman ini hanya bisa diakses oleh Admin.")
        return view_func(request, *args, **kwargs)
    return wrapper


def requester_or_admin_required(view_func):
    """Decorator that restricts access to requester or admin."""
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if not (_is_admin(request.user) or _is_requester(request.user)):
            return _forbidden_response(request, "Halaman ini hanya bisa diakses oleh Requester atau Admin.")
        return view_func(request, *args, **kwargs)
    return wrapper


def staff_or_admin_required(view_func):
    """Decorator that restricts access to staff or admin."""
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if not (_is_admin(request.user) or _is_staff_role(request.user)):
            return _forbidden_response(request, "Halaman ini hanya bisa diakses oleh Staff atau Admin.")
        return view_func(request, *args, **kwargs)
    return wrapper


def requester_staff_or_admin_required(view_func):
    """Decorator that restricts access to requester, staff, or admin."""
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if not _can_access_request_edit(request.user):
            return _forbidden_response(request, "Halaman ini hanya bisa diakses oleh Requester, Staff, atau Admin.")
        return view_func(request, *args, **kwargs)
    return wrapper


def executor_or_admin_required(view_func):
    """Decorator that restricts access to executor or admin."""
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if not (_is_admin(request.user) or _is_executor(request.user)):
            return _forbidden_response(request, "Halaman ini hanya bisa diakses oleh Executor atau Admin.")
        return view_func(request, *args, **kwargs)
    return wrapper


# --- Dashboard ---

@login_required
def dashboard(request):
    doc_qs = _filter_doc_requests_for_user(DocumentationRequest.objects.all(), request.user)
    maint_qs = _filter_maint_requests_for_user(MaintenanceRequest.objects.all(), request.user)

    # Jadwal tayang dapat dilihat oleh semua user yang login.
    jt_qs = JadwalTayang.objects.all()

    total_requests = doc_qs.count()
    total_maint = maint_qs.count()
    total_jt = jt_qs.count()

    # Per-status counts
    doc_todo = doc_qs.filter(status='TODO').count()
    doc_progress = doc_qs.filter(status='IN_PROGRESS').count()
    doc_done = doc_qs.filter(status='DONE').count()

    maint_todo = maint_qs.filter(status='TODO').count()
    maint_progress = maint_qs.filter(status='IN_PROGRESS').count()
    maint_done = maint_qs.filter(status='DONE').count()

    jt_todo = jt_qs.filter(status='BELUM_TAYANG').count()
    jt_progress = jt_qs.filter(status='SEDANG_TAYANG').count()
    jt_done = jt_qs.filter(status='SUDAH_TAKEOUT').count()

    # Recent items
    recent_docs = doc_qs.select_related(
        'brand', 'jenis_led', 'submitted_by'
    ).prefetch_related(
        'lokasi', 'materi_items'
    ).order_by('-created_at')[:5]
    recent_maints = maint_qs.select_related(
        'submitted_by', 'brand'
    ).prefetch_related('lokasi', 'materi_items').order_by('-created_at')[:5]
    recent_jt = jt_qs.select_related(
        'brand', 'jenis_led', 'submitted_by'
    ).prefetch_related('lokasi', 'materi_items').order_by('-created_at')[:5]

    return render(request, "products/dashboard.html", {
        "total_requests": total_requests,
        "total_maint": total_maint,
        "total_jt": total_jt,
        "doc_todo": doc_todo,
        "doc_progress": doc_progress,
        "doc_done": doc_done,
        "maint_todo": maint_todo,
        "maint_progress": maint_progress,
        "maint_done": maint_done,
        "jt_todo": jt_todo,
        "jt_progress": jt_progress,
        "jt_done": jt_done,
        "recent_docs": recent_docs,
        "recent_maints": recent_maints,
        "recent_jt": recent_jt,
        "can_create_requests": _is_requester(request.user) or _is_admin(request.user),
    })


# --- Documentation Request Views ---

@login_required
def doc_request_list(request):
    search_query = _get_search_query(request)
    requests = _filter_doc_requests_for_user(
        DocumentationRequest.objects.select_related(
        "brand", "jenis_led", "submitted_by"
    ).prefetch_related(
        "lokasi",
        "materi_items",
        "requirements",
        "view_photo",
        "jenis_kamera",
        "lokasi_assignments__lokasi",
        "lokasi_assignments__pelaksana",
    ).order_by("-id"),
        request.user,
    )
    if search_query:
        requests = requests.filter(
            _pk_search_q(search_query)
            | Q(brand__name__icontains=search_query)
            | Q(materi_items__nama_materi__icontains=search_query)
            | Q(lokasi__name__icontains=search_query)
            | Q(jenis_led__name__icontains=search_query)
            | Q(requirements__name__icontains=search_query)
            | Q(jenis_kamera__name__icontains=search_query)
            | Q(note__icontains=search_query)
            | Q(pic_pemohon__icontains=search_query)
            | Q(status__icontains=search_query)
            | Q(submitted_by__username__icontains=search_query)
            | Q(submitted_by__first_name__icontains=search_query)
            | Q(submitted_by__last_name__icontains=search_query)
        ).distinct()
    return render(request, "products/request_list.html", {
        "requests": requests,
        "all_dokumentators": _assignable_dokumentators_queryset(),
        "can_create_requests": _is_requester(request.user) or _is_admin(request.user),
        "can_edit_requests": _can_access_request_edit(request.user),
        "is_admin": _is_admin(request.user),
        "is_staff_role": _is_staff_role(request.user),
        "is_executor": _is_executor(request.user),
        "can_manage_pelaksana": _can_manage_service_pelaksana(request.user),
        "can_upload_proof": _can_upload_service_proof(request.user),
        **_search_context(request, "Cari brand, lokasi, PIC, requirement, kamera, atau user"),
    })


@requester_or_admin_required
def doc_request_create(request):
    form = DocumentationRequestForm(request.POST or None)
    materi_rows = _extract_materi_rows(request.POST) if request.method == "POST" else _initial_materi_rows()
    if request.method == "POST" and not materi_rows:
        form.add_error(None, "Isi minimal satu materi.")
    if request.method == "POST" and form.is_valid() and materi_rows:
        lokasi_list = list(form.cleaned_data["lokasi"])
        requirements = list(form.cleaned_data["requirements"])
        view_photo = list(form.cleaned_data["view_photo"])
        jenis_kamera = list(form.cleaned_data["jenis_kamera"])

        with transaction.atomic():
            for lokasi in lokasi_list:
                doc_req = DocumentationRequest.objects.create(
                    submitted_by=request.user,
                    brand=form.cleaned_data["brand"],
                    jenis_led=form.cleaned_data["jenis_led"],
                    tanggal=form.cleaned_data["tanggal"],
                    note=form.cleaned_data["note"],
                    pic_pemohon=form.cleaned_data["pic_pemohon"],
                )
                doc_req.lokasi.set([lokasi])
                doc_req.requirements.set(requirements)
                doc_req.view_photo.set(view_photo)
                doc_req.jenis_kamera.set(jenis_kamera)
                _sync_documentation_request_materi(doc_req, materi_rows)
                _create_edit_history(
                    user=request.user,
                    action="CREATE",
                    request_type=EditHistory.RequestType.DOC_REQUEST,
                    object_id=doc_req.id,
                    label=_doc_request_label(doc_req),
                    new_value=f"Request baru dibuat untuk lokasi {lokasi.name} dengan materi: {_materi_display_from_rows(materi_rows)}",
                )
        return redirect("doc_request_list")
    return render(request, "products/request_form.html", {
        "form": form,
        "title": "Create Documentation Request",
        "materi_rows": materi_rows,
    })


@requester_staff_or_admin_required
def doc_request_edit(request, pk):
    doc_request = get_object_or_404(
        DocumentationRequest.objects.select_related(
            "submitted_by", "brand", "jenis_led"
        ).prefetch_related(
            "lokasi",
            "materi_items",
            "requirements",
            "view_photo",
            "jenis_kamera",
        ),
        pk=pk,
    )
    if not _can_edit_request_record(request.user, doc_request.submitted_by):
        return _forbidden_response(request, "Anda tidak memiliki izin untuk mengedit request ini.")

    form = DocumentationRequestEditForm(request.POST or None, instance=doc_request)
    materi_rows = _extract_materi_rows(request.POST) if request.method == "POST" else _initial_materi_rows(doc_request)
    old_values = _doc_request_edit_snapshot(doc_request) if request.method == "POST" else None

    if request.method == "POST" and not materi_rows:
        form.add_error(None, "Isi minimal satu materi.")
    if request.method == "POST" and form.is_valid() and materi_rows:
        with transaction.atomic():
            doc_request = form.save(commit=False)
            doc_request.save()
            doc_request.lokasi.set([form.cleaned_data["lokasi"]])
            doc_request.requirements.set(form.cleaned_data["requirements"])
            doc_request.view_photo.set(form.cleaned_data["view_photo"])
            doc_request.jenis_kamera.set(form.cleaned_data["jenis_kamera"])
            _sync_documentation_request_materi(doc_request, materi_rows)
            doc_request.refresh_from_db()
            new_values = _doc_request_edit_snapshot(doc_request)
            label = _doc_request_label(doc_request)
            for field_name, old_value in old_values.items():
                new_value = new_values[field_name]
                if old_value != new_value:
                    _create_edit_history(
                        user=request.user,
                        action="UPDATE",
                        request_type=EditHistory.RequestType.DOC_REQUEST,
                        object_id=doc_request.id,
                        label=label,
                        field_name=field_name,
                        old_value=old_value,
                        new_value=new_value,
                    )
        return redirect("doc_request_detail", pk=doc_request.pk)

    return render(request, "products/request_form.html", {
        "form": form,
        "title": "Edit Documentation Request",
        "is_edit": True,
        "doc_request": doc_request,
        "materi_rows": materi_rows,
    })


@login_required
def doc_request_detail(request, pk):
    doc_request = get_object_or_404(
        DocumentationRequest.objects.select_related(
            "submitted_by", "brand", "jenis_led"
        ).prefetch_related(
            "lokasi",
            "materi_items",
            "requirements",
            "view_photo",
            "jenis_kamera",
            "lokasi_assignments__lokasi",
            "lokasi_assignments__pelaksana",
        ),
        pk=pk,
    )
    if not _filter_doc_requests_for_user(DocumentationRequest.objects.filter(pk=pk), request.user).exists():
        return _forbidden_response(request, "Anda tidak memiliki izin untuk melihat detail request ini.")

    proof_error = ""
    can_upload_proof = _can_upload_service_proof(request.user)
    materi_items = list(doc_request.materi_items.all())

    if request.method == "POST":
        if not can_upload_proof:
            return _forbidden_response(request, "Anda tidak memiliki izin untuk upload bukti kerja.")
        uploaded_files = [
            request.FILES.get(f"foto_bukti_kerja_{materi.id}")
            for materi in materi_items
        ]
        proof_error = next(
            (
                _validate_uploaded_image_size(uploaded_file)
                for uploaded_file in uploaded_files
                if _validate_uploaded_image_size(uploaded_file)
            ),
            "",
        )
        if not proof_error:
            old_status = doc_request.get_status_display()
            label = _doc_request_label(doc_request)
            changed_any = False
            for materi in materi_items:
                uploaded_file = request.FILES.get(f"foto_bukti_kerja_{materi.id}")
                if not uploaded_file:
                    continue
                changed_any = True
                old_file = _format_file_for_history(materi.foto_bukti_kerja)
                if materi.foto_bukti_kerja and materi.foto_bukti_kerja.storage.exists(materi.foto_bukti_kerja.name):
                    materi.foto_bukti_kerja.delete(save=False)
                materi.foto_bukti_kerja = uploaded_file
                materi.save(update_fields=["foto_bukti_kerja"])
                new_file = _format_file_for_history(materi.foto_bukti_kerja)
                _create_edit_history(
                    user=request.user,
                    action="UPDATE",
                    request_type=EditHistory.RequestType.DOC_REQUEST,
                    object_id=doc_request.id,
                    label=label,
                    field_name=f"Foto Bukti Kerja ({materi.nama_materi})",
                    old_value=old_file,
                    new_value=new_file,
                )
            if changed_any:
                _sync_parent_status_from_materi_proofs(doc_request, materi_items)
                doc_request.refresh_from_db(fields=["status"])
            if changed_any and old_status != doc_request.get_status_display():
                _create_edit_history(
                    user=request.user,
                    action="UPDATE",
                    request_type=EditHistory.RequestType.DOC_REQUEST,
                    object_id=doc_request.id,
                    label=label,
                    field_name="Status",
                    old_value=old_status,
                    new_value=doc_request.get_status_display(),
                )
            return redirect("doc_request_detail", pk=doc_request.pk)

    return render(request, "products/request_detail.html", {
        "request": doc_request,
        "proof_error": proof_error,
        "can_upload_proof": can_upload_proof,
        "can_edit_request": _can_edit_request_record(request.user, doc_request.submitted_by),
        "is_admin": _is_admin(request.user),
    })


@admin_required
def doc_request_delete(request, pk):
    doc_request = get_object_or_404(DocumentationRequest, pk=pk)
    if request.method == "POST":
        label = _doc_request_label(doc_request)
        _create_edit_history(
            user=request.user,
            action="DELETE",
            request_type=EditHistory.RequestType.DOC_REQUEST,
            object_id=pk,
            label=label,
            old_value=label,
            new_value="Dihapus",
        )
        doc_request.delete()
        return redirect("doc_request_list")
    return render(request, "products/request_delete.html", {"request_obj": doc_request})


@admin_required
def doc_request_update_status(request, pk):
    """AJAX-only endpoint to update doc request status."""
    if request.method == "POST":
        doc_request = get_object_or_404(DocumentationRequest, pk=pk)
        old_status = doc_request.get_status_display()
        new_status = request.POST.get("status", "")
        valid = [c[0] for c in DocumentationRequest.STATUS_CHOICES]
        if new_status in valid:
            doc_request.status = new_status
            doc_request.save(update_fields=["status"])
            new_label = doc_request.get_status_display()
            _create_edit_history(
                user=request.user,
                action="UPDATE",
                request_type=EditHistory.RequestType.DOC_REQUEST,
                object_id=pk,
                label=_doc_request_label(doc_request),
                field_name="Status",
                old_value=old_status,
                new_value=new_label,
            )
            return JsonResponse({"success": True, "status": new_status})
        return JsonResponse({"success": False, "error": "Invalid status"}, status=400)
    return HttpResponseForbidden("POST only.")


@staff_or_admin_required
def doc_request_update_lokasi_pelaksana(request, assignment_pk):
    """AJAX-only endpoint to update pelaksana for a request lokasi assignment."""
    if request.method == "POST":
        assignment = get_object_or_404(
            DocumentationRequestLokasiAssignment.objects.select_related(
                "documentation_request", "lokasi"
            ).prefetch_related("pelaksana"),
            pk=assignment_pk,
        )
        old_names = assignment.pelaksana_display()
        pelaksana_ids = request.POST.getlist("pelaksana[]")
        assignment.pelaksana.set(pelaksana_ids)
        new_names = assignment.pelaksana_display()
        if old_names != new_names:
            _create_edit_history(
                user=request.user,
                action="UPDATE",
                request_type=EditHistory.RequestType.DOC_REQUEST,
                object_id=assignment.documentation_request_id,
                label=_doc_request_label(assignment.documentation_request),
                field_name=f"Pelaksana ({assignment.lokasi.name})",
                old_value=old_names,
                new_value=new_names,
            )
        return JsonResponse({"success": True, "pelaksana_display": new_names})
    return HttpResponseForbidden("POST only.")


# --- AJAX endpoint: create Lokasi on-the-fly ---

@login_required
def ajax_create_lokasi(request):
    """Allow any logged-in user to create a new Lokasi via AJAX."""
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        if not name:
            return JsonResponse({"success": False, "error": "Name is required"}, status=400)
        obj, created = Lokasi.objects.get_or_create(name=name)
        return JsonResponse({"success": True, "id": obj.id, "name": obj.name})
    return HttpResponseForbidden("POST only.")


@login_required
def ajax_create_brand(request):
    """Allow any logged-in user to create a new Brand via AJAX."""
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        if not name:
            return JsonResponse({"success": False, "error": "Name is required"}, status=400)
        obj, created = BrandMateri.objects.get_or_create(name=name)
        return JsonResponse({"success": True, "id": obj.id, "name": obj.name})
    return HttpResponseForbidden("POST only.")

# --- Edit History (Admin Only) ---

@admin_required
def edit_history_list(request):
    search_query = _get_search_query(request)
    history_qs = EditHistory.objects.select_related("user").all()
    if search_query:
        history_qs = history_qs.filter(
            _pk_search_q(search_query)
            | Q(action__icontains=search_query)
            | Q(request_type__icontains=search_query)
            | Q(doc_request_label__icontains=search_query)
            | Q(field_name__icontains=search_query)
            | Q(old_value__icontains=search_query)
            | Q(new_value__icontains=search_query)
            | Q(user__username__icontains=search_query)
            | Q(user__first_name__icontains=search_query)
            | Q(user__last_name__icontains=search_query)
        ).distinct()
    paginator = Paginator(history_qs, 25)
    page = request.GET.get("page")
    history = paginator.get_page(page)
    return render(request, "products/edit_history.html", {
        "history": history,
        **_search_context(request, "Cari user, aksi, jenis request, field, atau isi perubahan"),
    })


@login_required
def notification_list(request):
    notifications = get_active_takeout_notifications()
    search_query = _get_search_query(request)
    if search_query:
        notifications = [
            notification
            for notification in notifications
            if _contains_search_value(
                search_query,
                notification["rule_name"],
                notification["urgency_label"],
                notification["offset_display"],
                notification["jadwal_label"],
                notification["lokasi_label"],
                notification["message"],
                notification["takeout_at_display"],
                notification["time_status"],
                notification["title"],
                notification["pending_materi_label"],
                notification["progress_label"],
                notification["jadwal_tayang_id"],
                notification["rule_id"],
            )
        ]
    paginator = Paginator(notifications, 20)
    page = request.GET.get("page")
    notification_page = paginator.get_page(page)
    urgent_count = sum(1 for notification in notifications if notification["urgency"] == TakeoutAlertRule.Urgency.URGENT)
    warning_count = len(notifications) - urgent_count
    return render(
        request,
        "products/notification_list.html",
        {
            "notifications": notification_page,
            "notification_total": len(notifications),
            "urgent_count": urgent_count,
            "warning_count": warning_count,
            **_search_context(request, "Cari rule, urgency, jadwal, lokasi, atau status waktu"),
        },
    )


@login_required
def notification_summary(request):
    notifications = get_active_takeout_notifications()
    preview_notifications = notifications[:5]
    html = render_to_string(
        "products/_notification_dropdown_items.html",
        {
            "notifications": preview_notifications,
            "notification_total": len(notifications),
        },
        request=request,
    )
    urgent_notifications = [
        _serialize_notification_for_json(notification)
        for notification in notifications
        if notification["urgency"] == TakeoutAlertRule.Urgency.URGENT
    ]
    return JsonResponse(
        {
            "count": len(notifications),
            "urgent_count": len(urgent_notifications),
            "html": html,
            "urgent_notifications": urgent_notifications,
        }
    )


@admin_required
def takeout_alert_rule_list(request):
    search_query = _get_search_query(request)
    rules = TakeoutAlertRule.objects.all()
    if search_query:
        rules = rules.filter(
            _pk_search_q(search_query)
            | Q(name__icontains=search_query)
            | Q(trigger_direction__icontains=search_query)
            | Q(offset_unit__icontains=search_query)
            | Q(urgency__icontains=search_query)
        ).distinct()
    return render(request, "products/takeout_alert_rule_list.html", {
        "rules": rules,
        **_search_context(request, "Cari nama rule, trigger, offset, atau urgency"),
    })


@admin_required
def takeout_alert_rule_create(request):
    form = TakeoutAlertRuleForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        if _is_ajax(request):
            return JsonResponse({"success": True})
        return redirect("takeout_alert_rule_list")
    template = "products/_takeout_alert_rule_form_modal.html"
    return render(request, template, {"form": form, "title": "Tambah Aturan Notifikasi"})


@admin_required
def takeout_alert_rule_edit(request, pk):
    rule = get_object_or_404(TakeoutAlertRule, pk=pk)
    form = TakeoutAlertRuleForm(request.POST or None, instance=rule)
    if request.method == "POST" and form.is_valid():
        form.save()
        if _is_ajax(request):
            return JsonResponse({"success": True})
        return redirect("takeout_alert_rule_list")
    template = "products/_takeout_alert_rule_form_modal.html"
    return render(request, template, {"form": form, "title": f"Edit Aturan: {rule.name}", "rule": rule})


@admin_required
def takeout_alert_rule_delete(request, pk):
    rule = get_object_or_404(TakeoutAlertRule, pk=pk)
    if request.method == "POST":
        rule.delete()
        if _is_ajax(request):
            return JsonResponse({"success": True})
        return redirect("takeout_alert_rule_list")
    return render(request, "products/_takeout_alert_rule_delete_modal.html", {"rule": rule})


# --- Master Data Views (Admin Only) ---

MASTER_DATA_REGISTRY = {
    "brand-materi": {"model": BrandMateri, "label": "Brand", "icon": "bi-tag"},
    "lokasi": {"model": Lokasi, "label": "Lokasi", "icon": "bi-geo-alt"},
    "dokumentator": {"model": Dokumentator, "label": "Dokumentator", "icon": "bi-person-video3"},
    "led-type": {"model": LEDType, "label": "Jenis Produk", "icon": "bi-lightbulb"},
    "requirement": {"model": Requirement, "label": "Requirement", "icon": "bi-check2-square"},
    "view-photo": {"model": ViewPhoto, "label": "View Photo", "icon": "bi-camera"},
    "camera-type": {"model": cameratype, "label": "Jenis Kamera", "icon": "bi-webcam"},
}


@admin_required
def master_data_list(request, slug):
    config = MASTER_DATA_REGISTRY.get(slug)
    if not config:
        return HttpResponseForbidden("Not found.")
    search_query = _get_search_query(request)
    items = config["model"].objects.all().order_by("name")
    if search_query:
        items = items.filter(
            _pk_search_q(search_query) | Q(name__icontains=search_query)
        ).distinct()
    return render(request, "products/master_data_list.html", {
        "items": items,
        "config": config,
        "slug": slug,
        "registry": MASTER_DATA_REGISTRY,
        **_search_context(request, f"Cari {config['label']}"),
    })


@admin_required
def master_data_create(request, slug):
    config = MASTER_DATA_REGISTRY.get(slug)
    if not config:
        return HttpResponseForbidden("Not found.")
    form = MasterDataForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        config["model"].objects.create(name=form.cleaned_data["name"])
        if _is_ajax(request):
            return JsonResponse({"success": True})
        return redirect("master_data_list", slug=slug)
    template = "products/_master_data_form_modal.html" if _is_ajax(request) else "products/master_data_list.html"
    ctx = {"form": form, "config": config, "slug": slug, "title": f"Add {config['label']}"}
    return render(request, template, ctx)


@admin_required
def master_data_edit(request, slug, pk):
    config = MASTER_DATA_REGISTRY.get(slug)
    if not config:
        return HttpResponseForbidden("Not found.")
    item = get_object_or_404(config["model"], pk=pk)
    form = MasterDataForm(request.POST or None, initial={"name": item.name})
    if request.method == "POST" and form.is_valid():
        item.name = form.cleaned_data["name"]
        item.save()
        if _is_ajax(request):
            return JsonResponse({"success": True})
        return redirect("master_data_list", slug=slug)
    template = "products/_master_data_form_modal.html" if _is_ajax(request) else "products/master_data_list.html"
    ctx = {"form": form, "config": config, "slug": slug, "title": f"Edit {config['label']}"}
    return render(request, template, ctx)


@admin_required
def master_data_delete(request, slug, pk):
    config = MASTER_DATA_REGISTRY.get(slug)
    if not config:
        return HttpResponseForbidden("Not found.")
    item = get_object_or_404(config["model"], pk=pk)
    if request.method == "POST":
        item.delete()
        if _is_ajax(request):
            return JsonResponse({"success": True})
        return redirect("master_data_list", slug=slug)
    template = "products/_master_data_delete_modal.html" if _is_ajax(request) else "products/master_data_list.html"
    return render(request, template, {"item": item, "config": config, "slug": slug})


@admin_required
def master_data_export(request, slug):
    """Export master data as Excel (.xlsx) with a single 'name' column."""
    config = MASTER_DATA_REGISTRY.get(slug)
    if not config:
        return HttpResponseForbidden("Not found.")

    openpyxl = _load_openpyxl()
    if openpyxl is None:
        return HttpResponse("Fitur export Excel belum tersedia di server ini.", status=503)

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = config["label"].replace("/", "-")[:31]  # Excel limits title to 31 chars

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write header
    ws["A1"] = "name"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = header_alignment
    ws["A1"].border = thin_border

    # Write data
    items = list(config["model"].objects.all().order_by("name"))
    for idx, item in enumerate(items, start=2):
        cell = ws.cell(row=idx, column=1, value=item.name)
        cell.border = thin_border

    # Auto-fit column width
    max_len = max((len(str(item.name)) for item in items), default=10) if items else 10
    ws.column_dimensions["A"].width = max(max_len + 4, 20)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{slug}.xlsx"'
    return response


@admin_required
def master_data_import_preview(request, slug):
    """Preview import: parse Excel and return JSON with new vs duplicate items."""
    config = MASTER_DATA_REGISTRY.get(slug)
    if not config:
        return JsonResponse({"success": False, "error": "Not found."}, status=404)

    openpyxl = _load_openpyxl()
    if openpyxl is None:
        return JsonResponse(
            {"success": False, "error": "Fitur import Excel belum tersedia di server ini."},
            status=503,
        )

    if request.method != "POST":
        # GET: render the import modal
        return render(request, "products/_master_data_import_modal.html", {
            "config": config, "slug": slug,
        })

    excel_file = request.FILES.get("excel_file")
    if not excel_file:
        return JsonResponse({"success": False, "error": "File Excel diperlukan."}, status=400)

    if not excel_file.name.endswith((".xlsx", ".xls")):
        return JsonResponse({"success": False, "error": "File harus berformat .xlsx"}, status=400)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_file.read()), read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return JsonResponse({"success": False, "error": "File kosong."}, status=400)

        # Find 'name' column
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        if "name" not in headers:
            return JsonResponse(
                {"success": False, "error": 'Kolom "name" tidak ditemukan di file Excel.'},
                status=400,
            )
        name_col_idx = headers.index("name")

        # Parse names
        existing_names = set(
            config["model"].objects.values_list("name", flat=True)
        )
        new_items = []
        duplicate_items = []

        for row in rows[1:]:
            if name_col_idx < len(row) and row[name_col_idx]:
                name = str(row[name_col_idx]).strip()
                if not name:
                    continue
                if name in existing_names:
                    duplicate_items.append(name)
                else:
                    if name not in [n for n in new_items]:  # avoid dupes in file
                        new_items.append(name)
                    else:
                        duplicate_items.append(name)

        wb.close()
        return JsonResponse({
            "success": True,
            "new_items": new_items,
            "duplicate_items": duplicate_items,
            "total": len(new_items) + len(duplicate_items),
        })
    except Exception as e:
        return JsonResponse({"success": False, "error": f"Gagal membaca file: {str(e)}"}, status=400)


@admin_required
def master_data_import_confirm(request, slug):
    """Confirm and execute import of the provided names."""
    config = MASTER_DATA_REGISTRY.get(slug)
    if not config:
        return JsonResponse({"success": False, "error": "Not found."}, status=404)

    if request.method != "POST":
        return HttpResponseForbidden("POST only.")

    try:
        body = json.loads(request.body)
        names = body.get("names", [])
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({"success": False, "error": "Invalid request."}, status=400)

    created_count = 0
    for name in names:
        name = str(name).strip()
        if not name:
            continue
        _, created = config["model"].objects.get_or_create(name=name)
        if created:
            created_count += 1

    return JsonResponse({
        "success": True,
        "created": created_count,
        "message": f"{created_count} data berhasil ditambahkan.",
    })


# --- Maintenance & Troubleshoot LED Views ---

@login_required
def maint_request_list(request):
    search_query = _get_search_query(request)
    requests_qs = _filter_maint_requests_for_user(
        MaintenanceRequest.objects.select_related(
        "submitted_by", "brand"
    ).prefetch_related("lokasi", "materi_items", "jenis_led", "pelaksana").order_by("-id")
        ,
        request.user,
    )
    if search_query:
        requests_qs = requests_qs.filter(
            _pk_search_q(search_query)
            | Q(brand__name__icontains=search_query)
            | Q(materi_items__nama_materi__icontains=search_query)
            | Q(lokasi__name__icontains=search_query)
            | Q(nama_pemohon__icontains=search_query)
            | Q(departement__icontains=search_query)
            | Q(deskripsi_pekerjaan__icontains=search_query)
            | Q(status__icontains=search_query)
            | Q(submitted_by__username__icontains=search_query)
            | Q(submitted_by__first_name__icontains=search_query)
            | Q(submitted_by__last_name__icontains=search_query)
            | Q(jenis_led__name__icontains=search_query)
            | Q(pelaksana__name__icontains=search_query)
        ).distinct()
    return render(request, "products/maint_request_list.html", {
        "requests": requests_qs,
        "all_dokumentators": _assignable_dokumentators_queryset(),
        "can_create_requests": _is_requester(request.user) or _is_admin(request.user),
        "can_edit_requests": _can_access_request_edit(request.user),
        "is_admin": _is_admin(request.user),
        "is_staff_role": _is_staff_role(request.user),
        "is_executor": _is_executor(request.user),
        "can_manage_pelaksana": _can_manage_service_pelaksana(request.user),
        "can_upload_proof": _can_upload_service_proof(request.user),
        **_search_context(request, "Cari brand, lokasi, jenis produk, pemohon, departement, atau dokumentator"),
    })


@requester_or_admin_required
def maint_request_create(request):
    form = MaintenanceRequestForm(request.POST or None, request.FILES or None)
    materi_rows = _extract_materi_rows(request.POST) if request.method == "POST" else _initial_materi_rows()
    if request.method == "POST" and not materi_rows:
        form.add_error(None, "Isi minimal satu materi.")
    if request.method == "POST" and form.is_valid() and materi_rows:
        lokasi_list = list(form.cleaned_data["lokasi"])
        jenis_led = list(form.cleaned_data["jenis_led"])
        foto_kerusakan = form.cleaned_data.get("foto_kerusakan")

        with transaction.atomic():
            for lokasi in lokasi_list:
                if foto_kerusakan and hasattr(foto_kerusakan, "seek"):
                    foto_kerusakan.seek(0)
                maint_req = MaintenanceRequest.objects.create(
                    submitted_by=request.user,
                    nama_pemohon=form.cleaned_data["nama_pemohon"],
                    departement=form.cleaned_data["departement"],
                    tanggal_permintaan=form.cleaned_data["tanggal_permintaan"],
                    tanggal_deadline=form.cleaned_data["tanggal_deadline"],
                    brand=form.cleaned_data["brand"],
                    deskripsi_pekerjaan=form.cleaned_data["deskripsi_pekerjaan"],
                    foto_kerusakan=foto_kerusakan,
                )
                maint_req.lokasi.set([lokasi])
                maint_req.jenis_led.set(jenis_led)
                _sync_maintenance_request_materi(maint_req, materi_rows)
        return redirect("maint_request_list")

    return render(request, "products/maint_request_form.html", {
        "form": form,
        "title": "Request Maintenance & Troubleshoot LED",
        "materi_rows": materi_rows,
    })


@requester_staff_or_admin_required
def maint_request_edit(request, pk):
    maint_request = get_object_or_404(
        MaintenanceRequest.objects.select_related(
            "submitted_by", "brand"
        ).prefetch_related("lokasi", "materi_items", "jenis_led", "pelaksana"),
        pk=pk,
    )
    if not _can_edit_request_record(request.user, maint_request.submitted_by):
        return _forbidden_response(request, "Anda tidak memiliki izin untuk mengedit request ini.")

    form = MaintenanceRequestEditForm(request.POST or None, request.FILES or None, instance=maint_request)
    materi_rows = _extract_materi_rows(request.POST) if request.method == "POST" else _initial_materi_rows(maint_request)
    if request.method == "POST" and not materi_rows:
        form.add_error(None, "Isi minimal satu materi.")
    if request.method == "POST" and form.is_valid() and materi_rows:
        maint_request = form.save(commit=False)
        maint_request.save()
        maint_request.lokasi.set([form.cleaned_data["lokasi"]])
        maint_request.jenis_led.set(form.cleaned_data["jenis_led"])
        _sync_maintenance_request_materi(maint_request, materi_rows)
        return redirect("maint_request_detail", pk=maint_request.pk)

    return render(request, "products/maint_request_form.html", {
        "form": form,
        "title": "Edit Maintenance & Troubleshoot LED Request",
        "is_edit": True,
        "maint_request": maint_request,
        "materi_rows": materi_rows,
    })


@login_required
def maint_request_detail(request, pk):
    maint_request = get_object_or_404(
        MaintenanceRequest.objects.select_related(
            "submitted_by", "brand"
        ).prefetch_related(
            "lokasi", "materi_items", "jenis_led", "pelaksana"
        ),
        pk=pk,
    )
    if not _filter_maint_requests_for_user(MaintenanceRequest.objects.filter(pk=pk), request.user).exists():
        return _forbidden_response(request, "Anda tidak memiliki izin untuk melihat detail request ini.")
    proof_error = ""
    can_upload_proof = _can_upload_service_proof(request.user)
    materi_items = list(maint_request.materi_items.all())

    if request.method == "POST":
        if not can_upload_proof:
            return _forbidden_response(request, "Anda tidak memiliki izin untuk upload bukti kerja.")
        uploaded_files = [
            request.FILES.get(f"foto_bukti_kerja_{materi.id}")
            for materi in materi_items
        ]
        proof_error = next(
            (
                _validate_uploaded_image_size(uploaded_file)
                for uploaded_file in uploaded_files
                if _validate_uploaded_image_size(uploaded_file)
            ),
            "",
        )
        if not proof_error:
            changed_any = False
            for materi in materi_items:
                uploaded_file = request.FILES.get(f"foto_bukti_kerja_{materi.id}")
                if not uploaded_file:
                    continue
                changed_any = True
                if materi.foto_bukti_kerja and materi.foto_bukti_kerja.storage.exists(materi.foto_bukti_kerja.name):
                    materi.foto_bukti_kerja.delete(save=False)
                materi.foto_bukti_kerja = uploaded_file
                materi.save(update_fields=["foto_bukti_kerja"])
            if changed_any:
                _sync_parent_status_from_materi_proofs(maint_request, materi_items)
            return redirect("maint_request_detail", pk=maint_request.pk)

    return render(request, "products/maint_request_detail.html", {
        "req": maint_request,
        "proof_error": proof_error,
        "can_upload_proof": can_upload_proof,
        "can_edit_request": _can_edit_request_record(request.user, maint_request.submitted_by),
        "is_admin": _is_admin(request.user),
    })


@admin_required
def maint_request_delete(request, pk):
    maint_request = get_object_or_404(MaintenanceRequest, pk=pk)
    if request.method == "POST":
        maint_request.delete()
        return redirect("maint_request_list")
    return render(request, "products/maint_request_delete.html", {"request_obj": maint_request})


@admin_required
def maint_request_update_status(request, pk):
    if request.method == "POST":
        maint_request = get_object_or_404(MaintenanceRequest, pk=pk)
        new_status = request.POST.get("status", "")
        valid = [c[0] for c in MaintenanceRequest.STATUS_CHOICES]
        if new_status in valid:
            maint_request.status = new_status
            maint_request.save(update_fields=["status"])
            return JsonResponse({"success": True, "status": new_status})
        return JsonResponse({"success": False, "error": "Invalid status"}, status=400)
    return HttpResponseForbidden("POST only.")


@staff_or_admin_required
def maint_request_update_pelaksana(request, pk):
    """AJAX-only endpoint to update pelaksana for a maintenance request."""
    if request.method == "POST":
        maint_request = get_object_or_404(MaintenanceRequest, pk=pk)
        pelaksana_ids = request.POST.getlist("pelaksana[]")
        maint_request.pelaksana.set(pelaksana_ids)
        return JsonResponse({"success": True})
    return HttpResponseForbidden("POST only.")


# --- Jadwal Tayang Views ---

@login_required
def jadwal_tayang_list(request):
    search_query = _get_search_query(request)
    qs = JadwalTayang.objects.select_related(
        "brand", "jenis_led", "submitted_by"
    ).prefetch_related(
        "lokasi",
        "pelaksana",
        "materi_items__foto_tayang_set",
        "materi_items__foto_takeout_set",
        "materi_items__bukti_playlist",
    ).all()
    if search_query:
        qs = qs.filter(
            _pk_search_q(search_query)
            | Q(brand__name__icontains=search_query)
            | Q(materi_items__nama_materi__icontains=search_query)
            | Q(lokasi__name__icontains=search_query)
            | Q(jenis_led__name__icontains=search_query)
            | Q(note_requester__icontains=search_query)
            | Q(note_executor__icontains=search_query)
            | Q(pic_pemohon__icontains=search_query)
            | Q(status__icontains=search_query)
            | Q(submitted_by__username__icontains=search_query)
            | Q(submitted_by__first_name__icontains=search_query)
            | Q(submitted_by__last_name__icontains=search_query)
            | Q(pelaksana__name__icontains=search_query)
        ).distinct()
    now = timezone.now()
    requests = list(qs)
    for req in requests:
        req.photo_status_info = _jadwal_tayang_photo_status_info(req, now)
    return render(request, "products/jadwal_tayang_list.html", {
        "requests": requests,
        "all_dokumentators": _assignable_dokumentators_queryset(),
        "is_requester": _is_requester(request.user),
        "is_executor": _is_executor(request.user),
        "is_admin": _is_admin(request.user),
        "can_manage_pelaksana": _can_manage_service_pelaksana(request.user),
        **_search_context(request, "Cari brand, lokasi, PIC, notes, pelaksana, atau user"),
    })


@login_required
def jadwal_tayang_report(request):
    search_query = _get_search_query(request)
    now = timezone.now()
    qs = JadwalTayang.objects.select_related(
        "brand", "jenis_led", "submitted_by"
    ).prefetch_related("lokasi", "pelaksana", "materi_items").filter(
        tanggal_tayang__lte=now,
        tanggal_takeout__gte=now,
    )

    if search_query:
        qs = qs.filter(
            _pk_search_q(search_query)
            | Q(brand__name__icontains=search_query)
            | Q(materi_items__nama_materi__icontains=search_query)
            | Q(lokasi__name__icontains=search_query)
            | Q(jenis_led__name__icontains=search_query)
            | Q(note_requester__icontains=search_query)
            | Q(note_executor__icontains=search_query)
            | Q(pic_pemohon__icontains=search_query)
            | Q(submitted_by__username__icontains=search_query)
            | Q(submitted_by__first_name__icontains=search_query)
            | Q(submitted_by__last_name__icontains=search_query)
            | Q(pelaksana__name__icontains=search_query)
        ).distinct()

    active_jadwal = list(qs.order_by("tanggal_takeout", "tanggal_tayang", "id"))
    report_groups = _group_jadwal_tayang_by_lokasi(active_jadwal)

    return render(request, "products/jadwal_tayang_report.html", {
        "report_groups": report_groups,
        "active_count": len(active_jadwal),
        "generated_at": now,
        **_search_context(request, "Cari brand, lokasi, PIC, pelaksana, atau user"),
    })


@requester_or_admin_required
def jadwal_tayang_create(request):
    form = JadwalTayangForm(request.POST or None)
    materi_rows = _extract_materi_rows(request.POST) if request.method == "POST" else _initial_materi_rows()
    if request.method == "POST" and not materi_rows:
        form.add_error(None, "Isi minimal satu materi.")
    if request.method == "POST" and form.is_valid() and materi_rows:
        lokasi_list = list(form.cleaned_data["lokasi"])
        with transaction.atomic():
            for lokasi in lokasi_list:
                jt = JadwalTayang(
                    submitted_by=request.user,
                    brand=form.cleaned_data["brand"],
                    jenis_led=form.cleaned_data["jenis_led"],
                    tanggal_tayang=form.cleaned_data["tanggal_tayang"],
                    tanggal_takeout=form.cleaned_data["tanggal_takeout"],
                    note_requester=form.cleaned_data["note_requester"],
                    pic_pemohon=form.cleaned_data["pic_pemohon"],
                )
                jt.save()
                jt.lokasi.set([lokasi])
                _sync_jadwal_tayang_materi(jt, materi_rows)
                _create_edit_history(
                    user=request.user,
                    action="CREATE",
                    request_type=EditHistory.RequestType.JADWAL_TAYANG,
                    object_id=jt.id,
                    label=_jadwal_tayang_label(jt),
                    new_value=f"Jadwal tayang baru dibuat untuk lokasi {lokasi.name} dengan materi: {_materi_display_from_rows(materi_rows)}",
                )
        return redirect("jadwal_tayang_list")
    return render(request, "products/jadwal_tayang_form.html", {
        "form": form,
        "title": "Buat Jadwal Tayang",
        "materi_rows": materi_rows,
    })


@requester_or_admin_required
def jadwal_tayang_edit(request, pk):
    jt = get_object_or_404(
        JadwalTayang.objects.select_related("brand", "jenis_led").prefetch_related("lokasi", "materi_items"),
        pk=pk,
    )
    form = JadwalTayangEditForm(request.POST or None, instance=jt)
    materi_rows = _extract_materi_rows(request.POST) if request.method == "POST" else _initial_materi_rows(jt)
    old_values = _jadwal_tayang_edit_snapshot(jt) if request.method == "POST" else None
    if request.method == "POST" and not materi_rows:
        form.add_error(None, "Isi minimal satu materi.")
    if request.method == "POST" and form.is_valid() and materi_rows:
        with transaction.atomic():
            jt = form.save(commit=False)
            jt.save()
            jt.lokasi.set([form.cleaned_data["lokasi"]])
            _sync_jadwal_tayang_materi(jt, materi_rows)
            jt.refresh_from_db()
            new_values = _jadwal_tayang_edit_snapshot(jt)
            label = _jadwal_tayang_label(jt)
            for field_name, old_value in old_values.items():
                new_value = new_values[field_name]
                if old_value != new_value:
                    _create_edit_history(
                        user=request.user,
                        action="UPDATE",
                        request_type=EditHistory.RequestType.JADWAL_TAYANG,
                        object_id=jt.id,
                        label=label,
                        field_name=field_name,
                        old_value=old_value,
                        new_value=new_value,
                    )
        return redirect("jadwal_tayang_detail", pk=jt.pk)
    return render(request, "products/jadwal_tayang_form.html", {
        "form": form,
        "title": "Edit Jadwal Tayang",
        "is_edit": True,
        "jadwal_tayang": jt,
        "materi_rows": materi_rows,
    })


@login_required
def jadwal_tayang_detail(request, pk):
    jt = get_object_or_404(
        JadwalTayang.objects.select_related(
            "submitted_by", "brand", "jenis_led"
        ).prefetch_related(
            "lokasi", "pelaksana",
            "materi_items__foto_tayang_set",
            "materi_items__foto_takeout_set",
            "materi_items__bukti_playlist",
        ),
        pk=pk,
    )
    jt.auto_update_status()
    jt = get_object_or_404(
        JadwalTayang.objects.select_related(
            "submitted_by", "brand", "jenis_led"
        ).prefetch_related(
            "lokasi", "pelaksana",
            "materi_items__foto_tayang_set",
            "materi_items__foto_takeout_set",
            "materi_items__bukti_playlist",
        ),
        pk=pk,
    )

    return render(request, "products/jadwal_tayang_detail.html", {
        "jt": jt,
        "is_requester": _is_requester(request.user),
        "is_executor": _is_executor(request.user),
        "is_admin": _is_admin(request.user),
    })


@admin_required
def jadwal_tayang_delete(request, pk):
    jt = get_object_or_404(
        JadwalTayang.objects.select_related("brand").prefetch_related("lokasi"),
        pk=pk,
    )
    if request.method == "POST":
        label = _jadwal_tayang_label(jt)
        _create_edit_history(
            user=request.user,
            action="DELETE",
            request_type=EditHistory.RequestType.JADWAL_TAYANG,
            object_id=pk,
            label=label,
            old_value=label,
            new_value="Dihapus",
        )
        jt.delete()
        return redirect("jadwal_tayang_list")
    return render(request, "products/jadwal_tayang_delete.html", {"request_obj": jt})


@executor_or_admin_required
def jadwal_tayang_update_status(request, pk):
    if request.method == "POST":
        jt = get_object_or_404(
            JadwalTayang.objects.select_related("brand").prefetch_related("lokasi"),
            pk=pk,
        )
        old_status = jt.get_status_display()
        new_status = request.POST.get("status", "")
        valid = [c[0] for c in JadwalTayang.STATUS_CHOICES]
        if new_status in valid:
            jt.status = new_status
            jt.save(update_fields=["status"])
            new_label = jt.get_status_display()
            if old_status != new_label:
                _create_edit_history(
                    user=request.user,
                    action="UPDATE",
                    request_type=EditHistory.RequestType.JADWAL_TAYANG,
                    object_id=pk,
                    label=_jadwal_tayang_label(jt),
                    field_name="Status",
                    old_value=old_status,
                    new_value=new_label,
                )
            return JsonResponse({"success": True, "status": new_status})
        return JsonResponse({"success": False, "error": "Invalid status"}, status=400)
    return HttpResponseForbidden("POST only.")


@staff_or_admin_required
def jadwal_tayang_update_pelaksana(request, pk):
    if request.method == "POST":
        jt = get_object_or_404(
            JadwalTayang.objects.select_related("brand").prefetch_related("lokasi"),
            pk=pk,
        )
        old_names = _joined_names(jt.pelaksana)
        pelaksana_ids = request.POST.getlist("pelaksana[]")
        jt.pelaksana.set(pelaksana_ids)
        new_names = _joined_names(jt.pelaksana)
        if old_names != new_names:
            _create_edit_history(
                user=request.user,
                action="UPDATE",
                request_type=EditHistory.RequestType.JADWAL_TAYANG,
                object_id=pk,
                label=_jadwal_tayang_label(jt),
                field_name="Pelaksana",
                old_value=old_names,
                new_value=new_names,
            )
        return JsonResponse({"success": True})
    return HttpResponseForbidden("POST only.")


@executor_or_admin_required
def _legacy_jadwal_tayang_upload_photos(request, pk):
    """Executor/Admin upload photos & notes for a Jadwal Tayang."""
    jt = get_object_or_404(
        JadwalTayang.objects.select_related("brand").prefetch_related(
            "lokasi",
            "materi_items__foto_tayang_set",
            "materi_items__foto_takeout_set",
            "materi_items__bukti_playlist",
        ),
        pk=pk,
    )

    if request.method == "POST":
        label = _jadwal_tayang_label(jt)
        old_status = jt.get_status_display()
        old_note_executor = jt.note_executor
        old_pelaksana = _joined_names(jt.pelaksana)
        materi_items = list(jt.materi_items.all())

        uploader_dokumentator, _ = _get_or_create_dokumentator_for_user(request.user)
        if uploader_dokumentator:
            jt.pelaksana.add(uploader_dokumentator)
            new_pelaksana = _joined_names(jt.pelaksana)
            if old_pelaksana != new_pelaksana:
                _create_edit_history(
                    user=request.user,
                    action="UPDATE",
                    request_type=EditHistory.RequestType.JADWAL_TAYANG,
                    object_id=pk,
                    label=label,
                    field_name="Pelaksana",
                    old_value=old_pelaksana,
                    new_value=new_pelaksana,
                )

        # Save executor notes
        note_executor = request.POST.get("note_executor", "").strip()
        if note_executor and note_executor != old_note_executor:
            jt.note_executor = note_executor
            jt.save(update_fields=["note_executor"])
            _create_edit_history(
                user=request.user,
                action="UPDATE",
                request_type=EditHistory.RequestType.JADWAL_TAYANG,
                object_id=pk,
                label=label,
                field_name="Notes Executor",
                old_value=old_note_executor or "-",
                new_value=note_executor,
            )

        # Foto Tayang (multiple)
        foto_tayang_files = request.FILES.getlist("foto_tayang")
        for f in foto_tayang_files:
            JadwalTayangFotoTayang.objects.create(jadwal_tayang=jt, foto=f)
        if foto_tayang_files:
            new_count = jt.foto_tayang_set.count()
            _create_edit_history(
                user=request.user,
                action="UPDATE",
                request_type=EditHistory.RequestType.JADWAL_TAYANG,
                object_id=pk,
                label=label,
                field_name="Foto Tayang",
                old_value=f"{initial_foto_tayang_count} foto",
                new_value=f"{new_count} foto (+{len(foto_tayang_files)} baru)",
            )

        # Bukti Playlist (pagi, siang, malam) — delete old files
        foto_pagi = request.FILES.get("foto_playlist_pagi")
        foto_siang = request.FILES.get("foto_playlist_siang")
        foto_malam = request.FILES.get("foto_playlist_malam")
        if foto_pagi or foto_siang or foto_malam:
            bukti, _ = JadwalTayangBuktiPlaylist.objects.get_or_create(jadwal_tayang=jt)
            before_slots = []
            if bukti.foto_pagi:
                before_slots.append("Pagi")
            if bukti.foto_siang:
                before_slots.append("Siang")
            if bukti.foto_malam:
                before_slots.append("Malam")
            if foto_pagi:
                if bukti.foto_pagi and bukti.foto_pagi.storage.exists(bukti.foto_pagi.name):
                    bukti.foto_pagi.delete(save=False)
                bukti.foto_pagi = foto_pagi
            if foto_siang:
                if bukti.foto_siang and bukti.foto_siang.storage.exists(bukti.foto_siang.name):
                    bukti.foto_siang.delete(save=False)
                bukti.foto_siang = foto_siang
            if foto_malam:
                if bukti.foto_malam and bukti.foto_malam.storage.exists(bukti.foto_malam.name):
                    bukti.foto_malam.delete(save=False)
                bukti.foto_malam = foto_malam
            bukti.save()
            after_slots = []
            if bukti.foto_pagi:
                after_slots.append("Pagi")
            if bukti.foto_siang:
                after_slots.append("Siang")
            if bukti.foto_malam:
                after_slots.append("Malam")
            _create_edit_history(
                user=request.user,
                action="UPDATE",
                request_type=EditHistory.RequestType.JADWAL_TAYANG,
                object_id=pk,
                label=label,
                field_name="Bukti Playlist",
                old_value=", ".join(before_slots) or "-",
                new_value=", ".join(after_slots) or "-",
            )

        # Foto Takeout (multiple)
        foto_takeout_files = request.FILES.getlist("foto_takeout")
        for f in foto_takeout_files:
            JadwalTayangFotoTakeout.objects.create(jadwal_tayang=jt, foto=f)
        if foto_takeout_files:
            new_count = jt.foto_takeout_set.count()
            _create_edit_history(
                user=request.user,
                action="UPDATE",
                request_type=EditHistory.RequestType.JADWAL_TAYANG,
                object_id=pk,
                label=label,
                field_name="Foto Takeout",
                old_value=f"{initial_foto_takeout_count} foto",
                new_value=f"{new_count} foto (+{len(foto_takeout_files)} baru)",
            )

        # Auto-update status based on photos
        jt.auto_update_status()
        jt.refresh_from_db(fields=["status"])
        new_status = jt.get_status_display()
        if old_status != new_status:
            _create_edit_history(
                user=request.user,
                action="UPDATE",
                request_type=EditHistory.RequestType.JADWAL_TAYANG,
                object_id=pk,
                label=label,
                field_name="Status",
                old_value=old_status,
                new_value=new_status,
            )

        return redirect("jadwal_tayang_detail", pk=pk)

    return redirect("jadwal_tayang_detail", pk=pk)


@executor_or_admin_required
def jadwal_tayang_upload_photos(request, pk):
    """Executor/Admin upload photos & notes for each materi in a Jadwal Tayang."""
    jt = get_object_or_404(
        JadwalTayang.objects.select_related("brand").prefetch_related(
            "lokasi",
            "materi_items__foto_tayang_set",
            "materi_items__foto_takeout_set",
            "materi_items__bukti_playlist",
        ),
        pk=pk,
    )

    if request.method != "POST":
        return redirect("jadwal_tayang_detail", pk=pk)

    label = _jadwal_tayang_label(jt)
    old_status = jt.get_status_display()
    old_note_executor = jt.note_executor
    old_pelaksana = _joined_names(jt.pelaksana)
    materi_items = list(jt.materi_items.all())
    old_materi_statuses = {
        materi.id: materi.get_status_display()
        for materi in materi_items
    }

    uploader_dokumentator, _ = _get_or_create_dokumentator_for_user(request.user)
    if uploader_dokumentator:
        jt.pelaksana.add(uploader_dokumentator)
        new_pelaksana = _joined_names(jt.pelaksana)
        if old_pelaksana != new_pelaksana:
            _create_edit_history(
                user=request.user,
                action="UPDATE",
                request_type=EditHistory.RequestType.JADWAL_TAYANG,
                object_id=pk,
                label=label,
                field_name="Pelaksana",
                old_value=old_pelaksana,
                new_value=new_pelaksana,
            )

    note_executor = request.POST.get("note_executor", "").strip()
    if note_executor and note_executor != old_note_executor:
        jt.note_executor = note_executor
        jt.save(update_fields=["note_executor"])
        _create_edit_history(
            user=request.user,
            action="UPDATE",
            request_type=EditHistory.RequestType.JADWAL_TAYANG,
            object_id=pk,
            label=label,
            field_name="Notes Executor",
            old_value=old_note_executor or "-",
            new_value=note_executor,
        )

    for materi in materi_items:
        initial_foto_tayang_count = materi.foto_tayang_set.count()
        initial_foto_takeout_count = materi.foto_takeout_set.count()

        foto_tayang_files = request.FILES.getlist(f"foto_tayang_{materi.id}")
        for uploaded_file in foto_tayang_files:
            JadwalTayangFotoTayang.objects.create(materi=materi, foto=uploaded_file)
        if foto_tayang_files:
            _create_edit_history(
                user=request.user,
                action="UPDATE",
                request_type=EditHistory.RequestType.JADWAL_TAYANG,
                object_id=pk,
                label=label,
                field_name=f"Foto Tayang ({materi.nama_materi})",
                old_value=f"{initial_foto_tayang_count} foto",
                new_value=f"{materi.foto_tayang_set.count()} foto (+{len(foto_tayang_files)} baru)",
            )

        foto_pagi = request.FILES.get(f"foto_playlist_pagi_{materi.id}")
        foto_siang = request.FILES.get(f"foto_playlist_siang_{materi.id}")
        foto_malam = request.FILES.get(f"foto_playlist_malam_{materi.id}")
        if foto_pagi or foto_siang or foto_malam:
            bukti, _ = JadwalTayangBuktiPlaylist.objects.get_or_create(materi=materi)
            before_slots = []
            if bukti.foto_pagi:
                before_slots.append("Pagi")
            if bukti.foto_siang:
                before_slots.append("Siang")
            if bukti.foto_malam:
                before_slots.append("Malam")
            if foto_pagi:
                if bukti.foto_pagi and bukti.foto_pagi.storage.exists(bukti.foto_pagi.name):
                    bukti.foto_pagi.delete(save=False)
                bukti.foto_pagi = foto_pagi
            if foto_siang:
                if bukti.foto_siang and bukti.foto_siang.storage.exists(bukti.foto_siang.name):
                    bukti.foto_siang.delete(save=False)
                bukti.foto_siang = foto_siang
            if foto_malam:
                if bukti.foto_malam and bukti.foto_malam.storage.exists(bukti.foto_malam.name):
                    bukti.foto_malam.delete(save=False)
                bukti.foto_malam = foto_malam
            bukti.save()
            after_slots = []
            if bukti.foto_pagi:
                after_slots.append("Pagi")
            if bukti.foto_siang:
                after_slots.append("Siang")
            if bukti.foto_malam:
                after_slots.append("Malam")
            _create_edit_history(
                user=request.user,
                action="UPDATE",
                request_type=EditHistory.RequestType.JADWAL_TAYANG,
                object_id=pk,
                label=label,
                field_name=f"Bukti Playlist ({materi.nama_materi})",
                old_value=", ".join(before_slots) or "-",
                new_value=", ".join(after_slots) or "-",
            )

        foto_takeout_files = request.FILES.getlist(f"foto_takeout_{materi.id}")
        for uploaded_file in foto_takeout_files:
            JadwalTayangFotoTakeout.objects.create(materi=materi, foto=uploaded_file)
        if foto_takeout_files:
            _create_edit_history(
                user=request.user,
                action="UPDATE",
                request_type=EditHistory.RequestType.JADWAL_TAYANG,
                object_id=pk,
                label=label,
                field_name=f"Foto Takeout ({materi.nama_materi})",
                old_value=f"{initial_foto_takeout_count} foto",
                new_value=f"{materi.foto_takeout_set.count()} foto (+{len(foto_takeout_files)} baru)",
            )

        materi.auto_update_status()
        materi.refresh_from_db(fields=["status"])
        new_materi_status = materi.get_status_display()
        if old_materi_statuses.get(materi.id) != new_materi_status:
            _create_edit_history(
                user=request.user,
                action="UPDATE",
                request_type=EditHistory.RequestType.JADWAL_TAYANG,
                object_id=pk,
                label=label,
                field_name=f"Status Materi ({materi.nama_materi})",
                old_value=old_materi_statuses.get(materi.id, "-"),
                new_value=new_materi_status,
            )

    jt.auto_update_status()
    jt.refresh_from_db(fields=["status"])
    new_status = jt.get_status_display()
    if old_status != new_status:
        _create_edit_history(
            user=request.user,
            action="UPDATE",
            request_type=EditHistory.RequestType.JADWAL_TAYANG,
            object_id=pk,
            label=label,
            field_name="Status",
            old_value=old_status,
            new_value=new_status,
        )

    return redirect("jadwal_tayang_detail", pk=pk)


# --- User Management (Admin Only) ---

@admin_required
def user_list(request):
    search_query = _get_search_query(request)
    users = _annotate_login_required_flag(
        User.objects.select_related("login_requirement").prefetch_related("groups")
    ).order_by("username")
    if search_query:
        users = users.filter(_user_search_filter(search_query)).distinct()
    return render(request, "products/user_list.html", {
        "users": users,
        **_search_context(request, "Cari username, nama, email, atau role"),
    })


@admin_required
def user_never_login_list(request):
    search_query = _get_search_query(request)
    users = _annotate_login_required_flag(
        User.objects.select_related("login_requirement").prefetch_related("groups")
    ).filter(
        login_required_flag=True,
    ).order_by("username")
    if search_query:
        users = users.filter(_user_search_filter(search_query)).distinct()
    login_required_count = users.count()
    never_login_count = users.filter(last_login__isnull=True).count()
    return render(request, "products/user_never_login_list.html", {
        "users": users,
        "login_required_count": login_required_count,
        "never_login_count": never_login_count,
        **_search_context(request, "Cari username, nama, email, atau role"),
    })


@admin_required
def user_update_login_requirement(request, pk):
    if request.method != "POST":
        return HttpResponseForbidden("POST only.")

    user_obj = get_object_or_404(User, pk=pk)
    requires_login = request.POST.get("requires_login") == "true"
    UserLoginRequirement.objects.update_or_create(
        user=user_obj,
        defaults={"requires_login": requires_login},
    )
    return JsonResponse({
        "success": True,
        "requires_login": requires_login,
    })


@admin_required
def user_create(request):
    form = UserForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("user_list")
    return render(request, "products/user_form.html", {"form": form, "title": "Create User"})


@admin_required
def user_edit(request, pk):
    user_obj = get_object_or_404(User, pk=pk)
    form = UserForm(request.POST or None, instance=user_obj)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("user_list")
    return render(request, "products/user_form.html", {"form": form, "title": f"Edit User: {user_obj.username}"})


@admin_required
def user_delete(request, pk):
    user_obj = get_object_or_404(User, pk=pk)
    if request.method == "POST":
        user_obj.delete()
        return redirect("user_list")
    return render(request, "products/user_delete.html", {"user_obj": user_obj})
